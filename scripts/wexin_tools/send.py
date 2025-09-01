import os
import sys
import openpyxl
from wxauto import WeChat
import time

def main(params):
    """
    发送微信消息
    
    参数:
        message_file: 消息内容文件
    """
    message_file = params['message_file']

    try:
        # 读取消息配置文件
        messages = read_messages(message_file)
        
        # 循环messages，调用SendMsg进行发送
        wx = WeChat()
        for item in messages:
            msgType = item['type']
            who = item['target']
            msg = item['content']
            delay = float(item['delay'])
            # 文本消息
            if msgType == 'text':
                wx.SendMsg(msg=msg, who=who)
                print(f"\n已发送: {item}")
                time.sleep(delay)
            # 文件：如图片、PDF、Word等
            elif msgType == 'file':
                wx.SendFiles(filepath=msg, who=who)
                print(f"\n已发送: {item}")
                time.sleep(delay)
        
        print(f"\n完成")

    except Exception as e:
        print(f"发生错误: {e}")


# 读取配置文件
def read_messages(message_file):
    try:
        # 获取message_file对应的文件，如"D:\message.xlsx"
        wb = openpyxl.load_workbook(message_file)
        ws = wb.active

        # 读取表格第二行的A列到D列，这里放到是字段定义
        column_keys = {}
        for col in range(1, 5):  # A=1, B=2, C=3, D=4
            cell_value = ws.cell(row=2, column=col).value
            column_letter = openpyxl.utils.get_column_letter(col)
            column_keys[column_letter] = cell_value

        # 读取表格内容，从第4行开始读取，直到空行为止
        messages = []
        row = 4
        while True:
            row_data = {}
            empty_row = True
            for col in range(1, 5):
                cell_value = ws.cell(row=row, column=col).value
                column_letter = openpyxl.utils.get_column_letter(col)
                key = column_keys[column_letter]
                row_data[key] = cell_value
                if cell_value is not None and str(cell_value).strip() != '':
                    empty_row = False
            if empty_row:
                break
            messages.append(row_data)
            row += 1

        print("字段定义:", column_keys)
        print("消息内容:")
        for msg in messages:
            print(msg)
        print(f"\n完成")
        return messages

    except Exception as e:
        print(f"发生错误: {e}")

# 参数转换
def parse_args_to_dict():
    """
    将命令行参数解析为字典
    格式要求: -key1 value1 -key2 value2 ...
    返回: 包含参数键值对的字典
    """
    print(sys.argv)
    args = sys.argv[1:]  # 跳过脚本名
    result = {}
    i = 0
    while i < len(args):
        if args[i].startswith('-'):
            key = args[i][1:]  # 去掉前导的'-'
            if i + 1 < len(args) and not args[i+1].startswith('-'):
                result[key] = args[i+1]
                i += 2
            else:
                result[key] = True  # 没有值的参数设为True
                i += 1
        else:
            i += 1
    return result


if __name__ == "__main__":
    # 参数转字典
    args_dict = parse_args_to_dict()
    print(args_dict)
    
    # 调用重命名函数
    main(args_dict)
